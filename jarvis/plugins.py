"""jarvis.plugins

On-demand plugin loader for heavy features.

Plugins are only imported when needed, keeping the base JARVIS
lightweight on 8GB systems.
"""

from __future__ import annotations

import importlib
import logging
import threading
import time
from pathlib import Path
from typing import Any, Callable

from jarvis import config

logger = logging.getLogger(__name__)


class PluginManager:
    """Manages on-demand loading of heavy feature plugins."""

    def __init__(self) -> None:
        self._loaded_plugins: dict[str, Any] = {}
        self._plugin_lock = threading.Lock()
        self._plugin_status: dict[str, str] = {}

    def load_plugin(self, name: str) -> Any | None:
        """Load a plugin by name on demand.
        
        Available plugins:
            - vision: Webcam, object detection, face recognition
            - ocr: Advanced OCR, document parsing
            - local_llm: Offline AI models (Phi-3, Gemma)
            - browser_agent: Selenium automation
            - file_ai: Advanced file operations
        """
        with self._plugin_lock:
            if name in self._loaded_plugins:
                return self._loaded_plugins[name]

            try:
                plugin = self._load_plugin_internal(name)
                if plugin:
                    self._loaded_plugins[name] = plugin
                    self._plugin_status[name] = "loaded"
                    logger.info(f"Plugin '{name}' loaded successfully")
                    return plugin
            except Exception as e:
                logger.exception(f"Failed to load plugin '{name}'")
                self._plugin_status[name] = f"error: {str(e)}"
                return None

        return None

    def _load_plugin_internal(self, name: str) -> Any | None:
        """Internal plugin loading logic."""
        if name == "vision":
            return self._load_vision_plugin()
        elif name == "ocr":
            return self._load_ocr_plugin()
        elif name == "local_llm":
            return self._load_local_llm_plugin()
        elif name == "browser_agent":
            return self._load_browser_agent_plugin()
        elif name == "file_ai":
            return self._load_file_ai_plugin()
        else:
            logger.error(f"Unknown plugin: {name}")
            return None

    def _load_vision_plugin(self) -> dict[str, Any] | None:
        """Load computer vision capabilities."""
        try:
            import cv2
            import numpy as np

            # Check if we can access webcam
            cap = cv2.VideoCapture(0)
            if not cap.isOpened():
                logger.warning("No webcam detected")
                return None
            cap.release()

            plugin = {
                "cv2": cv2,
                "np": np,
                "webcam_available": True,
                "functions": {
                    "detect_qr": self._make_qr_detector(cv2),
                    "read_webcam": self._make_webcam_reader(cv2),
                    "detect_motion": self._make_motion_detector(cv2, np),
                },
            }
            return plugin

        except ImportError:
            logger.error("Vision plugin requires opencv-python (pip install opencv-python)")
            return None

    def _make_qr_detector(self, cv2: Any) -> Callable[[], str]:
        """Create QR code detector function."""
        def detect_qr() -> str:
            try:
                cap = cv2.VideoCapture(0)
                if not cap.isOpened():
                    return "Cannot access webcam."

                # Try for 5 seconds to find a QR code
                start_time = time.time()
                detector = cv2.QRCodeDetector()

                while time.time() - start_time < 5:
                    ret, frame = cap.read()
                    if not ret:
                        continue

                    data, bbox, _ = detector.detectAndDecode(frame)
                    if data:
                        cap.release()
                        return f"QR Code detected: {data}"

                    time.sleep(0.1)

                cap.release()
                return "No QR code found within 5 seconds."

            except Exception as e:
                return f"QR detection failed: {str(e)}"

        return detect_qr

    def _make_webcam_reader(self, cv2: Any) -> Callable[[str], str]:
        """Create webcam capture function."""
        def read_webcam(save_path: str = "webcam_capture.jpg") -> str:
            try:
                cap = cv2.VideoCapture(0)
                if not cap.isOpened():
                    return "Cannot access webcam."

                ret, frame = cap.read()
                if ret:
                    cv2.imwrite(save_path, frame)
                    cap.release()
                    return f"Webcam image saved to {save_path}"
                else:
                    cap.release()
                    return "Failed to capture image."

            except Exception as e:
                return f"Webcam capture failed: {str(e)}"

        return read_webcam

    def _make_motion_detector(self, cv2: Any, np: Any) -> Callable[[int], str]:
        """Create motion detector function."""
        def detect_motion(duration: int = 10) -> str:
            try:
                cap = cv2.VideoCapture(0)
                if not cap.isOpened():
                    return "Cannot access webcam."

                # Initialize background
                ret, frame1 = cap.read()
                ret, frame2 = cap.read()

                if not ret:
                    cap.release()
                    return "Cannot read from webcam."

                start_time = time.time()
                motion_detected = False

                while time.time() - start_time < duration:
                    diff = cv2.absdiff(frame1, frame2)
                    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
                    blur = cv2.GaussianBlur(gray, (5, 5), 0)
                    _, thresh = cv2.threshold(blur, 20, 255, cv2.THRESH_BINARY)
                    dilated = cv2.dilate(thresh, None, iterations=3)
                    contours, _ = cv2.findContours(dilated, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

                    for contour in contours:
                        if cv2.contourArea(contour) > 1000:
                            motion_detected = True
                            break

                    if motion_detected:
                        break

                    frame1 = frame2
                    ret, frame2 = cap.read()
                    if not ret:
                        break

                    time.sleep(0.05)

                cap.release()
                return "Motion detected!" if motion_detected else "No motion detected."

            except Exception as e:
                return f"Motion detection failed: {str(e)}"

        return detect_motion

    def _load_ocr_plugin(self) -> dict[str, Any] | None:
        """Load OCR capabilities."""
        try:
            import pytesseract
            from PIL import Image

            # Test if tesseract is available
            try:
                pytesseract.get_tesseract_version()
            except Exception:
                logger.warning("Tesseract OCR not found in PATH")
                return None

            plugin = {
                "pytesseract": pytesseract,
                "Image": Image,
                "functions": {
                    "read_image": self._make_image_reader(pytesseract, Image),
                    "read_screen": self._make_screen_reader(pytesseract, Image),
                },
            }
            return plugin

        except ImportError:
            logger.error("OCR plugin requires pytesseract and Pillow")
            return None

    def _make_image_reader(self, pytesseract: Any, Image: Any) -> Callable[[str], str]:
        """Create image OCR function."""
        def read_image(image_path: str) -> str:
            try:
                img = Image.open(image_path)
                text = pytesseract.image_to_string(img)
                return text.strip() if text.strip() else "No text found in image."
            except Exception as e:
                return f"OCR failed: {str(e)}"

        return read_image

    def _make_screen_reader(self, pytesseract: Any, Image: Any) -> Callable[[], str]:
        """Create screen OCR function."""
        def read_screen() -> str:
            try:
                import pyautogui

                screenshot = pyautogui.screenshot()
                text = pytesseract.image_to_string(screenshot)
                return text.strip() if text.strip() else "No text detected on screen."
            except Exception as e:
                return f"Screen reading failed: {str(e)}"

        return read_screen

    def _load_local_llm_plugin(self) -> dict[str, Any] | None:
        """Load local LLM capabilities."""
        try:
            # Try llama-cpp-python for GGUF models
            from llama_cpp import Llama

            plugin = {
                "Llama": Llama,
                "loaded_models": {},
                "functions": {
                    "load_model": self._make_model_loader(Llama),
                    "generate": self._make_text_generator(Llama),
                },
            }
            return plugin

        except ImportError:
            logger.error("Local LLM plugin requires llama-cpp-python")
            return None

    def _make_model_loader(self, Llama: Any) -> Callable[[str, int], str]:
        """Create model loader function."""
        def load_model(model_path: str, n_ctx: int = 2048) -> str:
            try:
                # This would actually load the model (heavy operation)
                # For now, just verify the file exists
                if not Path(model_path).exists():
                    return f"Model file not found: {model_path}"

                # Model loading is expensive - only do it when actually needed
                return f"Model ready: {model_path} (load on first use)"

            except Exception as e:
                return f"Model loading failed: {str(e)}"

        return load_model

    def _make_text_generator(self, Llama: Any) -> Callable[[str, str], str]:
        """Create text generation function."""
        def generate(model_path: str, prompt: str, max_tokens: int = 256) -> str:
            try:
                # Load model (this is slow - ~5-10 seconds on your CPU)
                llm = Llama(model_path=model_path, n_ctx=2048, verbose=False)
                output = llm(prompt, max_tokens=max_tokens, stop=["</s>"])
                return output["choices"][0]["text"].strip()
            except Exception as e:
                return f"Generation failed: {str(e)}"

        return generate

    def _load_browser_agent_plugin(self) -> dict[str, Any] | None:
        """Load browser automation."""
        try:
            from selenium import webdriver
            from selenium.webdriver.common.by import By
            from selenium.webdriver.common.keys import Keys

            plugin = {
                "webdriver": webdriver,
                "By": By,
                "Keys": Keys,
                "active_driver": None,
                "functions": {
                    "open_browser": self._make_browser_opener(webdriver),
                    "navigate": self._make_navigator(),
                    "search": self._make_searcher(),
                    "click_element": self._make_element_clicker(),
                },
            }
            return plugin

        except ImportError:
            logger.error("Browser agent requires selenium (pip install selenium)")
            return None

    def _make_browser_opener(self, webdriver: Any) -> Callable[[], str]:
        """Create browser opener function."""
        def open_browser() -> str:
            try:
                # Use Chrome if available, else Firefox
                try:
                    from selenium.webdriver.chrome.service import Service as ChromeService
                    from selenium.webdriver.chrome.options import Options

                    options = Options()
                    options.add_argument("--start-maximized")
                    driver = webdriver.Chrome(options=options)
                except Exception:
                    driver = webdriver.Firefox()

                # Store in plugin for later use
                plugin = self._loaded_plugins.get("browser_agent")
                if plugin:
                    plugin["active_driver"] = driver

                return "Browser opened successfully."

            except Exception as e:
                return f"Failed to open browser: {str(e)}"

        return open_browser

    def _make_navigator(self) -> Callable[[str], str]:
        """Create navigation function."""
        def navigate(url: str) -> str:
            try:
                plugin = self._loaded_plugins.get("browser_agent")
                if not plugin or not plugin.get("active_driver"):
                    return "No active browser. Say 'open browser' first."

                driver = plugin["active_driver"]
                driver.get(url)
                return f"Navigated to {url}"

            except Exception as e:
                return f"Navigation failed: {str(e)}"

        return navigate

    def _make_searcher(self) -> Callable[[str], str]:
        """Create search function."""
        def search(query: str) -> str:
            try:
                plugin = self._loaded_plugins.get("browser_agent")
                if not plugin or not plugin.get("active_driver"):
                    return "No active browser. Say 'open browser' first."

                driver = plugin["active_driver"]
                driver.get(f"https://www.google.com/search?q={query}")
                return f"Searching for: {query}"

            except Exception as e:
                return f"Search failed: {str(e)}"

        return search

    def _make_element_clicker(self) -> Callable[[str], str]:
        """Create element clicker function."""
        def click_element(text: str) -> str:
            try:
                plugin = self._loaded_plugins.get("browser_agent")
                if not plugin or not plugin.get("active_driver"):
                    return "No active browser. Say 'open browser' first."

                driver = plugin["active_driver"]
                By = plugin["By"]

                # Try to find and click element containing text
                try:
                    element = driver.find_element(By.XPATH, f"//*[contains(text(), '{text}')]")
                    element.click()
                    return f"Clicked element with text: {text}"
                except Exception:
                    return f"Could not find element with text: {text}"

            except Exception as e:
                return f"Click failed: {str(e)}"

        return click_element

    def _load_file_ai_plugin(self) -> dict[str, Any] | None:
        """Load advanced file operations."""
        try:
            import PyPDF2
            from docx import Document
            import openpyxl

            plugin = {
                "PyPDF2": PyPDF2,
                "Document": Document,
                "openpyxl": openpyxl,
                "functions": {
                    "read_pdf": self._make_pdf_reader(PyPDF2),
                    "read_docx": self._make_docx_reader(Document),
                    "read_excel": self._make_excel_reader(openpyxl),
                    "bulk_rename": self._make_bulk_renamer(),
                },
            }
            return plugin

        except ImportError:
            logger.error("File AI requires PyPDF2, python-docx, openpyxl")
            return None

    def _make_pdf_reader(self, PyPDF2: Any) -> Callable[[str], str]:
        """Create PDF reader function."""
        def read_pdf(path: str, pages: int = 3) -> str:
            try:
                with open(path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    text = ""
                    for i, page in enumerate(reader.pages):
                        if i >= pages:
                            break
                        text += page.extract_text() + "\n"
                    return text.strip()[:2000] or "No text extracted from PDF."
            except Exception as e:
                return f"PDF reading failed: {str(e)}"

        return read_pdf

    def _make_docx_reader(self, Document: Any) -> Callable[[str], str]:
        """Create DOCX reader function."""
        def read_docx(path: str) -> str:
            try:
                doc = Document(path)
                text = "\n".join([para.text for para in doc.paragraphs])
                return text.strip()[:2000] or "No text in document."
            except Exception as e:
                return f"DOCX reading failed: {str(e)}"

        return read_docx

    def _make_excel_reader(self, openpyxl: Any) -> Callable[[str], str]:
        """Create Excel reader function."""
        def read_excel(path: str, sheet: int = 0) -> str:
            try:
                wb = openpyxl.load_workbook(path)
                sheet_names = wb.sheetnames
                if sheet < len(sheet_names):
                    ws = wb[sheet_names[sheet]]
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append(" | ".join([str(cell) for cell in row if cell]))
                    return "\n".join(data[:50])  # First 50 rows
                return "Sheet not found."
            except Exception as e:
                return f"Excel reading failed: {str(e)}"

        return read_excel

    def _make_bulk_renamer(self) -> Callable[[str, str, str], str]:
        """Create bulk file renamer function."""
        def bulk_rename(folder: str, pattern: str, replacement: str) -> str:
            try:
                from pathlib import Path

                folder_path = Path(folder)
                if not folder_path.exists():
                    return f"Folder not found: {folder}"

                renamed = 0
                for file in folder_path.iterdir():
                    if file.is_file() and pattern in file.name:
                        new_name = file.name.replace(pattern, replacement)
                        file.rename(file.parent / new_name)
                        renamed += 1

                return f"Renamed {renamed} files."

            except Exception as e:
                return f"Bulk rename failed: {str(e)}"

        return bulk_rename

    def get_plugin_status(self) -> dict[str, str]:
        """Get status of all plugins."""
        return {
            "loaded": list(self._loaded_plugins.keys()),
            "status": self._plugin_status,
        }

    def unload_plugin(self, name: str) -> bool:
        """Unload a plugin to free memory."""
        with self._plugin_lock:
            if name in self._loaded_plugins:
                del self._loaded_plugins[name]
                self._plugin_status[name] = "unloaded"
                logger.info(f"Plugin '{name}' unloaded")
                return True
            return False


# Global plugin manager
plugin_manager = PluginManager()


def load_plugin(name: str) -> str:
    """Tool: Load a plugin on demand."""
    plugin = plugin_manager.load_plugin(name)
    if plugin:
        return f"Plugin '{name}' loaded successfully."
    return f"Failed to load plugin '{name}'. Check logs for details."


def unload_plugin(name: str) -> str:
    """Tool: Unload a plugin to free memory."""
    success = plugin_manager.unload_plugin(name)
    return f"Plugin '{name}' unloaded." if success else f"Plugin '{name}' was not loaded."


def get_plugin_status() -> str:
    """Tool: Get plugin loading status."""
    status = plugin_manager.get_plugin_status()
    loaded = ", ".join(status["loaded"]) if status["loaded"] else "None"
    return f"Loaded plugins: {loaded}"


# Plugin tool registry
PLUGIN_REGISTRY = {
    "load_plugin": load_plugin,
    "unload_plugin": unload_plugin,
    "get_plugin_status": get_plugin_status,
}
